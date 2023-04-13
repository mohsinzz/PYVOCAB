from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
import time

from plyer import notification
from plyer.facades.notification import Notification

wb=load_workbook('vocabulary.xlsx')
ws=wb.active
for row in range(1,len(ws['A'])):
    for c in range(1,2):
        char=get_column_letter(c)
        notification.notify(title="Vocabulary",message=ws[char+str(row)].value,app_icon=None,timeout=20)
        time.sleep(10)
    
