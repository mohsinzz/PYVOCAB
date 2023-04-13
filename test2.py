#import the required modules
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
import time
from time import localtime
from plyer import notification
#loading the respective workbooks
wb1=load_workbook('vocabulary.xlsx')
ws1=wb1.active
wb2=load_workbook('vocabrevision.xlsx')
ws2=wb2.active
# for saturday and sunday we want the the words
#from the vocabrevision to be displayed
if localtime().tm_wday==5 or  localtime().tm_wday==6:
    while ws2['A2'].value!=None:
        notification.notify(title="Vocabulary Revision",message=ws2['A2'].value,app_icon='D:\Mohsin\pyproject\iconbook.ico',timeout=15)
        ws2.delete_rows(2)
        wb2.save('vocabrevesion.xlsx')
       time.sleep(30*75)
# On weekdays we want the words from the vocabulary file
# and get them appended for the revision in the vocabrevision
else:
    while ws1['A2'].value!=None:
        notification.notify(title="Vocabulary",message=ws1['A2'].value,app_icon='D:\Mohsin\pyproject\iconbook.ico',timeout=15)
        x=ws1['A2'].value
        ws2.append([x])
        ws1.delete_rows(2)
        wb1.save('vocabulary.xlsx')
        wb2.save('vocabrevision.xlsx')
        time.sleep(60*75)
